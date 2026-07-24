"""Experiment: can we create a cell-anchored comment on Google Sheets?

Tries several Drive API anchor formats on the course spreadsheet (cell T50
of sheet 4333K), then an XLSX round-trip via openpyxl on a NEW converted file.
"""
from __future__ import annotations

import io
import json
import os
from datetime import datetime, timezone

import gspread
import requests
from dotenv import load_dotenv
from oauth2client.service_account import ServiceAccountCredentials

load_dotenv()

CRED = os.getenv("CREDENTIALS_FILE", "credentials.json")
SHEET_ID = "1t2sAxKgXqS4yB7fTM_dj6JKOcW7kGjtmHAQXbdBWQC0"
WS_TITLE = "4333K"
TEST_ROW, TEST_COL = 50, 20  # T50 — far from grade cells
MARKER = f"[plagiarism-anchor-experiment {datetime.now(timezone.utc).isoformat()}]"

SCOPES = [
    "https://spreadsheets.google.com/feeds",
    "https://www.googleapis.com/auth/drive",
]


def main() -> None:
    creds = ServiceAccountCredentials.from_json_keyfile_name(CRED, SCOPES)
    token = creds.get_access_token().access_token
    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    gc = gspread.authorize(creds)
    ss = gc.open_by_key(SHEET_ID)
    ws = ss.worksheet(WS_TITLE)
    sheet_gid = ws.id
    print(f"worksheet={WS_TITLE!r} gid={sheet_gid} test_cell=T50")

    list_url = f"https://www.googleapis.com/drive/v3/files/{SHEET_ID}/comments"
    r = requests.get(
        list_url,
        headers=headers,
        params={
            "fields": "comments(id,content,anchor,quotedFileContent,createdTime)",
            "pageSize": 20,
        },
        timeout=30,
    )
    print("LIST status", r.status_code)
    comments = r.json().get("comments", []) if r.ok else []
    print(f"existing comments: {len(comments)}")
    for c in comments[:8]:
        content = (c.get("content") or "")[:80].encode("ascii", "replace").decode()
        print(
            "  id=",
            c.get("id"),
            "anchor=",
            c.get("anchor"),
            "content=",
            content,
        )

    # Match the format already stored on this spreadsheet by our notifier,
    # plus variants with real sheet gid / proprietary range id style.
    anchors = {
        "none": None,
        "like_existing_tab0": json.dumps(
            {
                "type": "workbook-range",
                "sheetsTabId": 0,
                "startRowIndex": TEST_ROW - 1,
                "endRowIndex": TEST_ROW,
                "startColumnIndex": TEST_COL - 1,
                "endColumnIndex": TEST_COL,
            }
        ),
        "like_existing_real_gid": json.dumps(
            {
                "type": "workbook-range",
                "sheetsTabId": sheet_gid,
                "startRowIndex": TEST_ROW - 1,
                "endRowIndex": TEST_ROW,
                "startColumnIndex": TEST_COL - 1,
                "endColumnIndex": TEST_COL,
            }
        ),
        "proprietary_range_id": json.dumps(
            {
                "type": "workbook-range",
                "uid": 0,
                "range": "999000111",
            }
        ),
        "context_html": "__CONTEXT__",
    }

    create_url = f"https://www.googleapis.com/drive/v3/files/{SHEET_ID}/comments"
    for name, anchor in anchors.items():
        body: dict = {"content": f"{MARKER} variant={name}"}
        if name == "context_html":
            body["context"] = {
                "type": "text/html",
                "value": f"cell T50 sheet {WS_TITLE}",
            }
        elif anchor is not None:
            body["anchor"] = anchor
        resp = requests.post(
            create_url,
            headers=headers,
            params={"fields": "id,content,anchor,htmlContent,quotedFileContent"},
            json=body,
            timeout=30,
        )
        created = resp.json() if resp.ok else {"error": resp.text[:300]}
        print(
            f"CREATE {name}: HTTP {resp.status_code} "
            f"id={created.get('id')} anchor={created.get('anchor')!r}"
        )

    try:
        from openpyxl import load_workbook
        from openpyxl.comments import Comment as XLComment
    except ImportError:
        print("openpyxl not installed — skipping xlsx experiment")
        print("DONE")
        return

    export_url = f"https://www.googleapis.com/drive/v3/files/{SHEET_ID}/export"
    er = requests.get(
        export_url,
        headers={"Authorization": f"Bearer {token}"},
        params={
            "mimeType": (
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        },
        timeout=120,
    )
    print("EXPORT xlsx", er.status_code, "bytes", len(er.content) if er.ok else er.text[:200])
    if not er.ok:
        print("DONE")
        return

    bio = io.BytesIO(er.content)
    wb = load_workbook(bio)
    if WS_TITLE not in wb.sheetnames:
        print("sheet missing in export:", wb.sheetnames[:10])
        print("DONE")
        return

    xws = wb[WS_TITLE]
    cell = xws.cell(row=TEST_ROW, column=TEST_COL)
    cell.value = cell.value or "anchor-test"
    cell.comment = XLComment(f"{MARKER} xlsx-comment", "lab-grader-bot")
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)

    meta = {
        "name": f"plagiarism-anchor-xlsx-test-{datetime.now().strftime('%H%M%S')}",
        "mimeType": "application/vnd.google-apps.spreadsheet",
    }
    files = {
        "data": ("metadata", json.dumps(meta), "application/json; charset=UTF-8"),
        "file": (
            "book.xlsx",
            out.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ),
    }
    up = requests.post(
        "https://www.googleapis.com/upload/drive/v3/files"
        "?uploadType=multipart&fields=id,name,webViewLink",
        headers={"Authorization": f"Bearer {token}"},
        files=files,
        timeout=180,
    )
    print("UPLOAD converted", up.status_code, up.text[:500])
    if not up.ok:
        print("DONE")
        return

    new_id = up.json()["id"]
    print("NEW FILE link=", up.json().get("webViewLink"))
    lr = requests.get(
        f"https://www.googleapis.com/drive/v3/files/{new_id}/comments",
        headers=headers,
        params={"fields": "comments(id,content,anchor)", "pageSize": 50},
        timeout=30,
    )
    print("NEW FILE Drive comments:", lr.status_code, json.dumps(lr.json(), ensure_ascii=False)[:1000])

    try:
        nss = gc.open_by_key(new_id)
        nws = nss.worksheet(WS_TITLE)
        note = nws.get_note("T50") if hasattr(nws, "get_note") else None
        print("NEW FILE T50 note=", repr(note))
    except Exception as e:
        print("gspread check failed", e)

    print("DONE")


if __name__ == "__main__":
    main()
